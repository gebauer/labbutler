"""LabSuit export profile: the known column layout and a workbook row iterator.

The LabSuit spreadsheet export has one sheet per former item type, each sharing a set of
core columns plus type-specific extra columns. This module knows that layout and yields
clean per-row dictionaries; it does no parsing or persistence of its own.
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# Sheets that are not item data.
SKIP_SHEETS = {"Import Instructions"}

# Control columns used by LabSuit's re-import, not item attributes.
CONTROL_COLUMNS = {"Delete? (y/n)", "In stock? (y/n)"}

# Core columns that map onto dedicated Item fields.
CORE_TO_FIELD = {
    "SERIAL_NUMBER": "legacy_serial",
    "NAME": "name",
    "CATALOG_NUMBER": "catalog_number",
    "CAS_NUMBER": "cas_number",
    "LOT_NUMBER": "lot_number",
    "BARCODE": "barcode",
}

# Core columns handled specially (parsed, or mapped to relations).
SPECIAL_COLUMNS = {
    "TAGS",
    "PRICE",
    "EXPIRATION_DATE",
    "LOCATION",
    "SUB_LOCATION",
    "SUB_LOCATION2",
    "SUPPLIER",
    "OWNER",
}

# Core columns with no dedicated field: kept verbatim in custom_fields (standard LabSuit
# fields, so they do NOT register lab custom-field-pool definitions).
CORE_PASSTHROUGH = {
    "MANUFACTURER",
    "SUPPLY_DATE",
    "COMMENTS",
    "URL1",
    "URL2",
    "PACK_SIZE",
    "AMOUNT_IN_STOCK",
}

# Everything in this union is a "core" column; anything else on a sheet is an extra
# column that feeds the lab-level custom-field pool.
CORE_COLUMNS = (
    set(CONTROL_COLUMNS)
    | set(CORE_TO_FIELD)
    | set(SPECIAL_COLUMNS)
    | set(CORE_PASSTHROUGH)
)


@dataclass
class LabSuitRow:
    sheet: str
    row_number: int  # 1-based row in the sheet, including the header
    values: dict[str, object]  # column header -> cell value (blanks dropped)
    extra_columns: list[str]  # non-core column headers present on this sheet


def iter_rows(path: str | Path) -> Iterator[LabSuitRow]:
    """Yield each data row of a LabSuit workbook as a :class:`LabSuitRow`."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            if sheet.title in SKIP_SHEETS:
                continue
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            headers = [str(h).strip() if h is not None else "" for h in header]
            if "SERIAL_NUMBER" not in headers:
                # Not an item sheet (e.g. a stray notes tab).
                continue
            extra = [h for h in headers if h and h not in CORE_COLUMNS]
            for offset, raw_row in enumerate(rows, start=2):
                values = {
                    headers[i]: cell
                    for i, cell in enumerate(raw_row)
                    if i < len(headers) and headers[i] and cell not in (None, "")
                }
                if not values:
                    continue
                yield LabSuitRow(sheet.title, offset, values, extra)
    finally:
        workbook.close()
