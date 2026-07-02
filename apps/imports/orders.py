"""LabSuit orders export profile: the known column layout and a workbook row iterator.

The LabSuit *orders* export has one sheet per calendar year (``Year - 2026`` …), each
sharing the same fixed column layout of one procurement request per row. This module
knows that layout and yields clean per-row dictionaries; it does no parsing or
persistence of its own (mirroring :mod:`apps.imports.labsuit` for inventory).
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from pathlib import Path

import openpyxl

# Sheets that are not order data.
SKIP_SHEETS = {"Import Instructions"}

# The column that both identifies a real order sheet and titles the request.
NAME_COLUMN = "ITEM_NAME"


@dataclass
class OrderRow:
    sheet: str
    row_number: int  # 1-based row in the sheet, including the header
    values: dict[str, object]  # column header -> cell value (blanks dropped)


def iter_rows(path: str | Path) -> Iterator[OrderRow]:
    """Yield each data row of a LabSuit orders workbook as an :class:`OrderRow`."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            if sheet.title in SKIP_SHEETS:
                continue
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if not header:
                continue
            headers = [str(h).strip() if h is not None else "" for h in header]
            if NAME_COLUMN not in headers:
                # Not an order sheet (e.g. a stray notes tab).
                continue
            for offset, raw_row in enumerate(rows, start=2):
                values = {
                    headers[i]: cell
                    for i, cell in enumerate(raw_row)
                    if i < len(headers) and headers[i] and cell not in (None, "")
                }
                if not values:
                    continue
                yield OrderRow(sheet.title, offset, values)
    finally:
        workbook.close()
