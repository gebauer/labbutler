"""Generic spreadsheet importer: map arbitrary columns onto Item fields, then reuse the
LabSuit pipeline's plan + commit.

Unlike :mod:`apps.imports.labsuit`, which knows a fixed column layout, this module is
driven by a per-column *mapping* the user supplies through the import wizard. Reading and
plan-building here are pure (no DB writes); persistence goes through
:func:`apps.imports.service.commit`, which allocates a fresh frozen ``human_id`` for the
serial-less rows a generic source produces.
"""

from __future__ import annotations

from collections.abc import Iterator
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import openpyxl
from django.utils.text import slugify

from . import parsers
from .service import ImportPlan, ParsedRow

# --- Mapping targets ------------------------------------------------------------------

IGNORE = "ignore"
CUSTOM = "custom"

# Targets copied verbatim onto an Item field. value -> (model field, human label).
DIRECT_FIELDS: dict[str, tuple[str, str]] = {
    "name": ("name", "Name"),
    "legacy_serial": ("legacy_serial", "Legacy ID / serial"),
    "catalog_number": ("catalog_number", "Catalog number"),
    "cas_number": ("cas_number", "CAS number"),
    "lot_number": ("lot_number", "Lot number"),
    "barcode": ("barcode", "Barcode"),
    "wgk": ("wgk", "WGK (water hazard class)"),
    "storage_class": ("storage_class", "Storage class"),
}

# Targets needing parsing or a related lookup. value -> human label.
SPECIAL_FIELDS: dict[str, str] = {
    "price": "Price (amount + currency)",
    "expiration_date": "Expiration date",
    "location": "Location (one level per column)",
    "tag": "Tag(s) (comma-separated)",
    "vendor": "Vendor / supplier",
    "owner": "Owner (email)",
}

# Ordered (value, label) options offered for every column in the wizard.
TARGET_CHOICES: list[tuple[str, str]] = (
    [(IGNORE, "— ignore —")]
    + [(key, label) for key, (_, label) in DIRECT_FIELDS.items()]
    + list(SPECIAL_FIELDS.items())
    + [(CUSTOM, "Custom field (named after the column)")]
)

_VALID_TARGETS = {value for value, _ in TARGET_CHOICES}


# --- Reading --------------------------------------------------------------------------


@dataclass
class WorkbookColumns:
    """Headers and a few sample rows from one sheet, for building the mapping UI."""

    sheet_names: list[str]
    sheet: str
    headers: list[str]
    preview_rows: list[list[str]]


def _coerce(value: object) -> object:
    """Strip strings; leave dates/numbers as-is for downstream parsers."""
    if isinstance(value, str):
        return value.strip()
    return value


def sheet_names(path: str | Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_columns(
    path: str | Path, sheet: str | None = None, *, preview: int = 3
) -> WorkbookColumns:
    """Read the header row and up to ``preview`` data rows from ``sheet`` (or the first)."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        names = list(workbook.sheetnames)
        chosen = sheet if sheet in names else names[0]
        worksheet = workbook[chosen]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None) or ()
        headers = [str(h).strip() if h is not None else "" for h in header]
        preview_rows = []
        for raw in rows:
            if len(preview_rows) >= preview:
                break
            preview_rows.append(
                ["" if cell is None else str(_coerce(cell)) for cell in raw][: len(headers)]
            )
        return WorkbookColumns(names, chosen, headers, preview_rows)
    finally:
        workbook.close()


def iter_rows(path: str | Path, sheet: str) -> Iterator[dict[str, object]]:
    """Yield each data row of ``sheet`` as a ``{header: value}`` dict (blanks dropped)."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return
        headers = [str(h).strip() if h is not None else "" for h in header]
        for raw in rows:
            values = {
                headers[i]: _coerce(cell)
                for i, cell in enumerate(raw)
                if i < len(headers) and headers[i] and cell not in (None, "")
            }
            if values:
                yield values
    finally:
        workbook.close()


# --- Mapping ---------------------------------------------------------------------------


# Header substrings -> a sensible default target, to pre-fill the wizard.
_GUESSES: list[tuple[tuple[str, ...], str]] = [
    (("serial", "legacy"), "legacy_serial"),
    (("cas",), "cas_number"),
    (("catalog", "cat no", "cat.", "art"), "catalog_number"),
    (("lot",), "lot_number"),
    (("barcode", "ean"), "barcode"),
    (("name", "product", "reagent", "compound", "description"), "name"),
    (("price", "cost", "amount eur", "value"), "price"),
    (("expir", "expiry", "exp date", "best before", "haltbar"), "expiration_date"),
    (("location", "storage", "room", "shelf", "fridge", "freezer"), "location"),
    (("vendor", "supplier", "manufacturer", "hersteller", "lieferant"), "vendor"),
    (("owner", "responsible", "user"), "owner"),
    (("tag", "category", "type", "class"), "tag"),
    (("wgk",), "wgk"),
]


def guess_target(header: str) -> str:
    """Best-guess mapping target for a column header (falls back to ignore)."""
    lowered = header.lower()
    for needles, target in _GUESSES:
        if any(needle in lowered for needle in needles):
            return target
    return IGNORE


def validate_mapping(mapping: dict[str, str]) -> list[str]:
    """Return human-readable problems with a column->target mapping (empty = valid)."""
    errors = []
    targets = set(mapping.values())
    if "name" not in targets:
        errors.append("Map one column to “Name” — every item needs a name.")
    unknown = {t for t in targets if t not in _VALID_TARGETS}
    if unknown:
        errors.append(f"Unknown target(s): {', '.join(sorted(unknown))}.")
    return errors


def _apply(row: ParsedRow, target: str, header: str, value: object) -> None:
    """Mutate ``row`` by applying one mapped column value to its target."""
    if target in DIRECT_FIELDS:
        field_name, _ = DIRECT_FIELDS[target]
        text = str(value).strip()
        row.fields[field_name] = text
        if field_name == "legacy_serial":
            row.legacy_serial = text
        return

    if target == "price":
        price = parsers.parse_price(value)
        if price.warning:
            row.warnings.append(price.warning)
        if price.amount is not None:
            row.fields["price_amount"] = price.amount
            row.fields["price_currency"] = price.currency
        return

    if target == "expiration_date":
        candidate = value.date() if isinstance(value, datetime) else value
        parsed = parsers.parse_european_date(candidate)
        if parsed is None:
            row.warnings.append(f"unparseable expiration date: {value!r}")
        else:
            row.fields["expiration_date"] = parsed
        return

    if target == "location":
        level = parsers.parse_location_part(value)
        if level is not None:
            row.location_path.append(level)
        return

    if target == "tag":
        for name in str(value).split(","):
            name = name.strip()
            if name and name not in row.tag_names:
                row.tag_names.append(name)
        return

    if target == "vendor":
        row.vendor_name = str(value).strip()
        return

    if target == "owner":
        owner = str(value).strip()
        if "@" in owner:
            row.owner_email = owner.lower()
        elif owner:
            row.warnings.append(f"owner is not an email: {owner!r}")
        return

    if target == CUSTOM:
        key = slugify(header).replace("-", "_") or "field"
        row.custom_fields[key] = str(value).strip()
        row.field_pool_keys.append((key, header))


# Upper bound on data rows per import: the whole plan is held in memory, so an
# unbounded (or maliciously padded) spreadsheet must not be able to exhaust the worker.
MAX_IMPORT_ROWS = 20_000


class ImportTooLarge(ValueError):
    """Raised when a spreadsheet holds more data rows than one import may process."""


def build_generic_plan(
    rows: Iterator[dict[str, object]],
    mapping: dict[str, str],
    *,
    max_rows: int = MAX_IMPORT_ROWS,
) -> ImportPlan:
    """Build a dry-run :class:`ImportPlan` from raw rows and a column->target mapping.

    Pure: no DB access. Rows carry no ``human_id`` (generic sources have no stable
    identifier), so :func:`apps.imports.service.commit` allocates a fresh frozen ID and
    always creates — generic imports never silently update existing items. Raises
    :class:`ImportTooLarge` beyond ``max_rows`` data rows.
    """
    plan = ImportPlan()
    for offset, values in enumerate(rows, start=2):  # row 1 is the header
        if offset - 1 > max_rows:
            raise ImportTooLarge(
                f"the sheet has more than {max_rows} data rows; split the file and retry"
            )
        row = ParsedRow(sheet="", row_number=offset)
        for header, raw_value in values.items():
            target = mapping.get(header, IGNORE)
            if target == IGNORE:
                continue
            _apply(row, target, header, raw_value)
        if not str(row.fields.get("name", "")).strip():
            row.errors.append("missing name")
        plan.rows.append(row)
    return plan


def plan_from_file(path: str | Path, sheet: str, mapping: dict[str, str]) -> ImportPlan:
    """Convenience wrapper: read ``sheet`` from ``path`` and build a plan for ``mapping``."""
    return build_generic_plan(iter_rows(path, sheet), mapping)
